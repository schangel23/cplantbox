#!/usr/bin/env python3
"""Extract per-rank empirical leaf-collar emergence (t_col) from Vidal 2021
SupData3 sheet ``1.tem_tcol_des``, average M40+M52 per rank, and append
``t_col_emp_Cd`` to ``dart/coupling/data/phase_III_per_rank_LEAF.json``.

The empirical t_col anchors MultiPhaseStemGrowth::calcLengthPerPhytomer's
Phase II transition (Vidal 2021 directly observed leaf-collar emergence
event), decoupling stem timing from the leaf-curve C¹ rescaling. See
`Literature/Chapter 1/Concepts/CPBOXBALENOCOUPLING/PLAN_VIDAL_TCOL_STEM_ANCHOR_2026-05-01.md`.

Usage::

    cpbenv/bin/python -m dart.coupling.scripts.vidal_tcol_extract \\
        --xlsx /home/lukas/Downloads/plaa072_suppl_supplementary_data_3.xlsx \\
        --json dart/coupling/data/phase_III_per_rank_LEAF.json \\
        [--apply]

Without ``--apply``: prints the per-rank averaged values (dry run).
With ``--apply``: writes ``ranks[].t_col_emp_Cd`` and a
``vidal_tcol_provenance`` block under ``_meta`` in-place.
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import statistics
import sys
from pathlib import Path

import openpyxl

SHEET = "1.tem_tcol_des"
COL_CULTIVAR = "cultivar"
COL_NF = "nf"
COL_TCOL = "tcol"
CULTIVARS = ("M40", "M52")


def load_tcol_table(xlsx: Path) -> dict[int, dict[str, float]]:
    """Return ``{rank: {cultivar: tcol}}`` from sheet 1.tem_tcol_des."""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[SHEET]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx_cult = headers.index(COL_CULTIVAR)
    idx_nf = headers.index(COL_NF)
    idx_tcol = headers.index(COL_TCOL)
    table: dict[int, dict[str, float]] = {}
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        cult = row[idx_cult]
        if cult not in CULTIVARS:
            continue
        rank = int(row[idx_nf])
        tcol = float(row[idx_tcol])
        table.setdefault(rank, {})[cult] = tcol
    return table


def average_per_rank(table: dict[int, dict[str, float]]) -> dict[int, dict]:
    """Return per-rank ``{rank: {avg, m40, m52, abs_dev}}`` averaged over M40+M52."""
    out: dict[int, dict] = {}
    for rank, by_cult in sorted(table.items()):
        vals = [by_cult[c] for c in CULTIVARS if c in by_cult]
        if len(vals) < 2:
            continue
        avg = statistics.mean(vals)
        out[rank] = {
            "avg": avg,
            "M40": by_cult.get("M40"),
            "M52": by_cult.get("M52"),
            "abs_dev_Cd": abs(by_cult["M40"] - by_cult["M52"]) / 2.0,
        }
    return out


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    h.update(path.read_bytes())
    return h.hexdigest()


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n", 1)[0])
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--json", required=True, type=Path)
    ap.add_argument("--apply", action="store_true",
                    help="write t_col_emp_Cd into the JSON in-place "
                         "(default: dry-run, print only)")
    args = ap.parse_args()

    table = load_tcol_table(args.xlsx)
    per_rank = average_per_rank(table)

    print(f"Vidal SupData3 sheet '{SHEET}' — M40+M52 averaged t_col per rank")
    print(f"{'rank':>4}  {'M40':>6}  {'M52':>6}  {'avg':>7}  {'±dev':>6}")
    for rank, info in sorted(per_rank.items()):
        print(f"  {rank:>2}  {info['M40']:>6.0f}  {info['M52']:>6.0f}"
              f"  {info['avg']:>7.1f}  {info['abs_dev_Cd']:>6.1f}")

    if not args.apply:
        print("\n(dry-run; pass --apply to write into JSON)")
        return 0

    with args.json.open() as f:
        data = json.load(f)

    # Write per-rank t_col_emp_Cd into ranks[] entries that match by `n`.
    # Skip gated ranks (1-3) — basal_zero_ranks short-circuits in
    # MultiPhaseStemGrowth before the empirical-anchor read, so writing
    # there is harmless but pointless. Skip ranks not in our JSON
    # (e.g. SupData3 rank 17/18 above ba2188fd's 15-leaf cap).
    n_written = 0
    for rank_entry in data["ranks"]:
        n = int(rank_entry["n"])
        if n in per_rank and not rank_entry.get("gated", False):
            info = per_rank[n]
            rank_entry["t_col_emp_Cd"] = round(info["avg"], 1)
            note = rank_entry.get("note", "")
            sep = " " if note and not note.endswith(" ") else ""
            rank_entry["note"] = (
                f"{note}{sep}t_col empirical {info['avg']:.1f} Cd "
                f"(M40+M52 avg, ±{info['abs_dev_Cd']:.1f} Cd from cultivar split)."
            )
            n_written += 1

    data["_meta"]["vidal_tcol_provenance"] = {
        "source_xlsx": args.xlsx.name,
        "source_sha256": sha256_of(args.xlsx),
        "fitter_script": "dart/coupling/scripts/vidal_tcol_extract.py",
        "applied_at": dt.date.today().isoformat(),
        "model_form_ref": "Vidal 2021 §Methods t_col observation (SupData3 sheet 1.tem_tcol_des)",
        "cultivars_used": list(CULTIVARS),
        "averaging_rule": "arithmetic mean of M40 and M52 per rank",
        "n_ranks_written": n_written,
        "rank_range_written": [
            min(per_rank.keys()), max(per_rank.keys())
        ],
        "consumer": "MultiPhaseStemGrowth::calcLengthPerPhytomer reads "
                    "LeafRandomParameter::t_col_emp_Cd to anchor stem "
                    "Phase II → directly on observed leaf-collar emergence, "
                    "decoupled from leaf-curve fits.",
    }

    with args.json.open("w") as f:
        json.dump(data, f, indent=2)
        f.write("\n")
    print(f"\n--apply: wrote t_col_emp_Cd for {n_written} ranks; "
          f"updated _meta.vidal_tcol_provenance.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
